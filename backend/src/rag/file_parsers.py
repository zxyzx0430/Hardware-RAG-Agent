"""
文件解析器模块。

支持格式：
  - PDF → Markdown（通过 PyMuPDF/Docling）
  - XLSX/XLS → Markdown 表格
  - CSV → Markdown 表格
  - JSON → 格式化输出
"""

import csv
import json
import logging
import os
import tempfile
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


class BaseParser(ABC):
    """文件解析器基类，兼容 LangChain DocumentLoader 接口。"""

    @abstractmethod
    def parse(self, file_path: Path) -> str:
        """解析文件，返回纯文本内容。"""
        ...

    def parse_from_bytes(self, data: bytes) -> str:
        """从字节流解析文件（默认实现：写入临时文件后调用 parse）。

        子类可覆盖此方法以提供更高效的内存解析。
        """
        suffix = getattr(self, "_default_suffix", ".bin")
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        try:
            tmp.write(data)
            tmp.close()
            return self.parse(Path(tmp.name))
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    def parse_from_string(self, raw: str) -> str:
        """从字符串解析（默认实现：转字节后调用 parse_from_bytes）。"""
        return self.parse_from_bytes(raw.encode("utf-8"))


class PdfParser(BaseParser):
    """使用 PyMuPDF 解析 PDF 为文本（fallback 到纯文本提取）。"""

    _default_suffix = ".pdf"

    def parse(self, file_path: Path) -> str:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(file_path))
            parts: list[str] = []
            for page_num in range(doc.page_count):
                page = doc[page_num]
                parts.append(page.get_text("text"))
            doc.close()
            return "\n\n".join(parts)
        except ImportError:
            logger.warning("PyMuPDF 未安装，PDF 解析不可用")
            return ""
        except Exception as e:
            logger.warning(f"PDF 解析失败 {file_path.name}: {e}")
            return ""

    def parse_from_bytes(self, data: bytes) -> str:
        """直接从字节解析 PDF，避免临时文件。"""
        try:
            import fitz
            doc = fitz.open(stream=data, filetype="pdf")
            parts: list[str] = []
            for page_num in range(doc.page_count):
                page = doc[page_num]
                parts.append(page.get_text("text"))
            doc.close()
            return "\n\n".join(parts)
        except ImportError:
            logger.warning("PyMuPDF 未安装，PDF 解析不可用")
            return ""
        except Exception as e:
            logger.warning(f"PDF 字节流解析失败: {e}")
            return ""


class XlsxParser(BaseParser):
    """使用 openpyxl 解析 XLSX/XLS 文件，转为 Markdown 表格。"""

    _default_suffix = ".xlsx"

    def parse(self, file_path: Path) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## {sheet_name}\n")

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 第一行作为表头
            header = [str(c) if c is not None else "" for c in rows[0]]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join("---" for _ in header) + " |")

            for row in rows[1:]:
                cells = [str(c).replace("\n", " ") if c is not None else "" for c in row]
                # 补齐列数
                while len(cells) < len(header):
                    cells.append("")
                parts.append("| " + " | ".join(cells[:len(header)]) + " |")

            parts.append("")  # 空行分隔

        wb.close()
        return "\n".join(parts)

    def parse_from_bytes(self, data: bytes) -> str:
        """直接从字节解析 XLSX，避免临时文件。"""
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## {sheet_name}\n")

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = [str(c) if c is not None else "" for c in rows[0]]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join("---" for _ in header) + " |")

            for row in rows[1:]:
                cells = [str(c).replace("\n", " ") if c is not None else "" for c in row]
                while len(cells) < len(header):
                    cells.append("")
                parts.append("| " + " | ".join(cells[:len(header)]) + " |")

            parts.append("")

        wb.close()
        return "\n".join(parts)


# ExcelParser 作为 XlsxParser 的别名（兼容 common.py/kb_routes.py 调用）
ExcelParser = XlsxParser


class CsvParser(BaseParser):
    """使用 csv 标准库解析 CSV 文件，转为 Markdown 表格。"""

    _default_suffix = ".csv"

    def parse(self, file_path: Path) -> str:
        import io

        try:
            import chardet
            raw = file_path.read_bytes()
            detected = chardet.detect(raw)
            encoding = detected.get("encoding", "utf-8") or "utf-8"
            text = raw.decode(encoding, errors="replace")
        except Exception:
            logger.warning("编码检测失败，使用 UTF-8 回退: %s", file_path.name)
            text = file_path.read_text(encoding="utf-8", errors="replace")

        return self._parse_text(text)

    def parse_from_bytes(self, data: bytes) -> str:
        """从字节解析 CSV，自动检测编码。"""
        try:
            import chardet
            detected = chardet.detect(data)
            encoding = detected.get("encoding", "utf-8") or "utf-8"
            text = data.decode(encoding, errors="replace")
        except Exception:
            text = data.decode("utf-8", errors="replace")
        return self._parse_text(text)

    def parse_from_string(self, raw: str) -> str:
        """从字符串解析 CSV。"""
        return self._parse_text(raw)

    def _parse_text(self, text: str) -> str:
        import io

        # 自动检测分隔符
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(text[:4096])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        rows = list(reader)

        if not rows:
            return ""

        parts: list[str] = []
        header = rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")

        for row in rows[1:]:
            cells = [c.replace("\n", " ") for c in row]
            while len(cells) < len(header):
                cells.append("")
            parts.append("| " + " | ".join(cells[:len(header)]) + " |")

        return "\n".join(parts)


class JsonParser(BaseParser):
    """解析 JSON 文件，格式化输出。"""

    _default_suffix = ".json"

    def parse(self, file_path: Path) -> str:
        raw = file_path.read_text(encoding="utf-8", errors="replace")
        return self._format_json(raw)

    def parse_from_string(self, raw: str) -> str:
        """从字符串解析 JSON（用于 chat 附件）。"""
        return self._format_json(raw)

    def parse_from_bytes(self, data: bytes) -> str:
        """从字节解析 JSON。"""
        return self._format_json(data.decode("utf-8", errors="replace"))

    def _format_json(self, raw: str) -> str:
        """格式化 JSON 字符串为可读文本。"""
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as e:
            return f"JSON 解析失败: {e}"

        # 如果是列表，逐项格式化
        if isinstance(data, list):
            parts: list[str] = []
            for i, item in enumerate(data):
                parts.append(f"### 条目 {i + 1}\n")
                parts.append(self._format_value(item))
                parts.append("")
            return "\n".join(parts)

        # 如果是字典，按键值格式化
        return self._format_value(data)

    def _format_value(self, data, indent: int = 0) -> str:
        """递归格式化 JSON 值。"""
        prefix = "  " * indent
        if isinstance(data, dict):
            if not data:
                return f"{prefix}(空对象)"
            lines = []
            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    lines.append(f"{prefix}- **{key}**:")
                    lines.append(self._format_value(value, indent + 1))
                else:
                    lines.append(f"{prefix}- **{key}**: {value}")
            return "\n".join(lines)
        elif isinstance(data, list):
            if not data:
                return f"{prefix}(空列表)"
            lines = []
            for item in data:
                if isinstance(item, (dict, list)):
                    lines.append(self._format_value(item, indent + 1))
                else:
                    lines.append(f"{prefix}- {item}")
            return "\n".join(lines)
        else:
            return f"{prefix}{data}"
